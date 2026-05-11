from fastapi import FastAPI, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse
import mysql.connector
from mysql.connector import Error
import bcrypt
import jwt
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from typing import Optional
from pydantic import BaseModel

app = FastAPI(title="Order Punching System")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Config ────────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "root",   # ← change this
    "database": "order_punch"
}
SECRET_KEY = "fresh-order-punch"  # ← change this
TOKEN_EXPIRE_HOURS = 8

# ── DB helper ─────────────────────────────────────────────────────────────────
def get_db():
    conn = mysql.connector.connect(**DB_CONFIG)
    try:
        yield conn
    finally:
        conn.close()

# ── Auth ──────────────────────────────────────────────────────────────────────
bearer_scheme = HTTPBearer()

def create_token(user_id: int, role: str) -> str:
    payload = {
        "sub": str(user_id),
        "role": role,
        "exp": datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm="HS256")

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(bearer_scheme)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=["HS256"])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def require_admin(token=Depends(verify_token)):
    if token.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return token

# ── Pydantic models ───────────────────────────────────────────────────────────
class LoginRequest(BaseModel):
    email: str
    password: str

class DistributorCreate(BaseModel):
    code: str
    name: str

class ProductCreate(BaseModel):
    code: str
    name: str

class OrderEntry(BaseModel):
    dist_id: int
    product_id: int
    quantity: int
    order_date: Optional[str] = None

class BulkOrderEntry(BaseModel):
    entries: list[OrderEntry]
    order_date: Optional[str] = None

class UserDistributorAssignment(BaseModel):
    dist_ids: list[int]

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.post("/auth/login")
def login(req: LoginRequest, db=Depends(get_db)):
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT * FROM users WHERE email=%s AND is_active=1", (req.email,))
    user = cur.fetchone()
    if not user or not bcrypt.checkpw(req.password.encode(), user["password_hash"].encode()):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_token(user["id"], user["role"])
    return {"token": token, "name": user["name"], "role": user["role"]}

# ── Distributor routes ────────────────────────────────────────────────────────
@app.get("/distributors")
def list_distributors(db=Depends(get_db), token=Depends(verify_token)):
    cur = db.cursor(dictionary=True)
    if token.get("role") == "admin":
        cur.execute("SELECT * FROM distributors WHERE is_active=1 ORDER BY name")
    else:
        user_id = int(token["sub"])
        cur.execute("""
            SELECT d.* FROM distributors d
            JOIN user_distributor_map m ON d.id = m.dist_id
            WHERE d.is_active=1 AND m.user_id=%s
            ORDER BY d.name
        """, (user_id,))
    return cur.fetchall()

@app.post("/distributors")
def add_distributor(data: DistributorCreate, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    try:
        cur.execute("INSERT INTO distributors (code, name) VALUES (%s, %s)", (data.code, data.name))
        db.commit()
        return {"id": cur.lastrowid, "message": "Distributor added"}
    except Error as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/distributors/{dist_id}")
def update_distributor(dist_id: int, data: DistributorCreate, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("UPDATE distributors SET code=%s, name=%s WHERE id=%s", (data.code, data.name, dist_id))
    db.commit()
    return {"message": "Updated"}

@app.delete("/distributors/{dist_id}")
def delete_distributor(dist_id: int, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("UPDATE distributors SET is_active=0 WHERE id=%s", (dist_id,))
    db.commit()
    return {"message": "Deactivated"}

# ── Product routes ────────────────────────────────────────────────────────────
@app.get("/products")
def list_products(db=Depends(get_db), token=Depends(verify_token)):
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT * FROM products WHERE is_active=1 ORDER BY name")
    return cur.fetchall()

@app.post("/products")
def add_product(data: ProductCreate, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    try:
        cur.execute("INSERT INTO products (code, name) VALUES (%s, %s)", (data.code, data.name))
        db.commit()
        return {"id": cur.lastrowid, "message": "Product added"}
    except Error as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.put("/products/{product_id}")
def update_product(product_id: int, data: ProductCreate, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("UPDATE products SET code=%s, name=%s WHERE id=%s", (data.code, data.name, product_id))
    db.commit()
    return {"message": "Updated"}

@app.delete("/products/{product_id}")
def delete_product(product_id: int, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("UPDATE products SET is_active=0 WHERE id=%s", (product_id,))
    db.commit()
    return {"message": "Deactivated"}

# ── Orders routes ─────────────────────────────────────────────────────────────
@app.get("/orders")
def get_orders(order_date: Optional[str] = None, db=Depends(get_db), token=Depends(verify_token)):
    cur = db.cursor(dictionary=True)
    if order_date:
        cur.execute("""
            SELECT o.*, d.name as dist_name, d.code as dist_code,
                   p.name as product_name, p.code as product_code
            FROM orders o
            JOIN distributors d ON o.dist_id = d.id
            JOIN products p ON o.product_id = p.id
            WHERE DATE(o.order_date) = %s
        """, (order_date,))
    else:
        cur.execute("""
            SELECT o.*, d.name as dist_name, d.code as dist_code,
                   p.name as product_name, p.code as product_code
            FROM orders o
            JOIN distributors d ON o.dist_id = d.id
            JOIN products p ON o.product_id = p.id
            WHERE DATE(o.order_date) = CURDATE()
        """)
    return cur.fetchall()

@app.post("/orders/bulk")
def save_orders(data: BulkOrderEntry, db=Depends(get_db), token=Depends(verify_token)):
    cur = db.cursor()
    user_id = int(token["sub"])
    role = token.get("role")
    order_date = data.order_date or datetime.now().strftime("%Y-%m-%d")

    if role != "admin":
        cur.execute("SELECT dist_id FROM user_distributor_map WHERE user_id=%s", (user_id,))
        allowed = {r[0] for r in cur.fetchall()}
        for entry in data.entries:
            if entry.quantity > 0 and entry.dist_id not in allowed:
                raise HTTPException(status_code=403, detail="Not authorized for this distributor")

    saved = 0
    for entry in data.entries:
        if entry.quantity <= 0:
            continue
        cur.execute("""
            SELECT quantity FROM orders
            WHERE dist_id=%s AND product_id=%s AND DATE(order_date)=%s
        """, (entry.dist_id, entry.product_id, order_date))
        existing = cur.fetchone()
        old_qty = existing[0] if existing else None
        action = "updated" if existing else "created"

        if old_qty == entry.quantity:
            saved += 1
            continue

        cur.execute("""
            INSERT INTO orders (dist_id, product_id, quantity, entered_by, order_date)
            VALUES (%s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE quantity=%s, entered_by=%s, updated_at=NOW()
        """, (entry.dist_id, entry.product_id, entry.quantity, user_id, order_date,
              entry.quantity, user_id))

        cur.execute("""
            INSERT INTO order_history
                (dist_id, product_id, order_date, old_quantity, new_quantity, changed_by, action)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (entry.dist_id, entry.product_id, order_date,
              old_qty, entry.quantity, user_id, action))
        saved += 1
    db.commit()
    return {"message": f"{saved} entries saved", "date": order_date}

@app.delete("/orders/clear")
def clear_orders(order_date: str, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("DELETE FROM orders WHERE DATE(order_date) = %s", (order_date,))
    db.commit()
    return {"message": f"Orders cleared for {order_date}"}

# ── Export ────────────────────────────────────────────────────────────────────
@app.get("/orders/export")
def export_orders(order_date: Optional[str] = None, db=Depends(get_db), token=Depends(require_admin)):
    cur = db.cursor(dictionary=True)
    date_filter = order_date or datetime.now().strftime("%Y-%m-%d")

    cur.execute("SELECT * FROM distributors WHERE is_active=1 ORDER BY name")
    distributors = cur.fetchall()

    cur.execute("SELECT * FROM products WHERE is_active=1 ORDER BY name")
    products = cur.fetchall()

    cur.execute("""
        SELECT dist_id, product_id, quantity FROM orders
        WHERE DATE(order_date) = %s
    """, (date_filter,))
    orders = {(r["dist_id"], r["product_id"]): r["quantity"] for r in cur.fetchall()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orders {date_filter}"

    # Styles
    header_fill = PatternFill("solid", fgColor="1B4F72")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sku_fill = PatternFill("solid", fgColor="2E86AB")
    sku_font = Font(bold=True, color="FFFFFF", size=10)
    total_fill = PatternFill("solid", fgColor="F0F3F4")
    total_font = Font(bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7")
    )

    # Header row: SKU col + one col per distributor + Total
    ws.cell(1, 1, "SKU / Distributor →").fill = header_fill
    ws.cell(1, 1).font = header_font
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).border = thin
    ws.column_dimensions["A"].width = 28

    col_map = {}
    for i, dist in enumerate(distributors, start=2):
        c = ws.cell(1, i, f"{dist['code']}\n{dist['name']}")
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin
        ws.column_dimensions[c.column_letter].width = 16
        col_map[dist["id"]] = i

    total_col = len(distributors) + 2
    tc = ws.cell(1, total_col, "Total")
    tc.fill = header_fill
    tc.font = header_font
    tc.alignment = center
    tc.border = thin
    ws.column_dimensions[tc.column_letter].width = 10
    ws.row_dimensions[1].height = 36

    # Data rows
    grand_total_col = {c: 0 for c in col_map.values()}
    for row_idx, product in enumerate(products, start=2):
        rc = ws.cell(row_idx, 1, f"[{product['code']}] {product['name']}")
        rc.fill = sku_fill
        rc.font = sku_font
        rc.alignment = Alignment(vertical="center")
        rc.border = thin

        row_total = 0
        for dist in distributors:
            qty = orders.get((dist["id"], product["id"]), 0)
            col = col_map[dist["id"]]
            cell = ws.cell(row_idx, col, qty if qty else "")
            cell.alignment = center
            cell.border = thin
            if qty:
                cell.fill = PatternFill("solid", fgColor="EAF2FF")
            row_total += qty
            grand_total_col[col] = grand_total_col.get(col, 0) + qty

        tc = ws.cell(row_idx, total_col, row_total if row_total else "")
        tc.fill = total_fill
        tc.font = total_font
        tc.alignment = center
        tc.border = thin
        ws.row_dimensions[row_idx].height = 20

    # Grand total row
    total_row = len(products) + 2
    gt = ws.cell(total_row, 1, "Grand Total")
    gt.fill = total_fill
    gt.font = total_font
    gt.border = thin
    grand_total = 0
    for col, val in grand_total_col.items():
        c = ws.cell(total_row, col, val if val else "")
        c.fill = total_fill
        c.font = total_font
        c.alignment = center
        c.border = thin
        grand_total += val
    gtc = ws.cell(total_row, total_col, grand_total)
    gtc.fill = PatternFill("solid", fgColor="D5E8D4")
    gtc.font = Font(bold=True, size=10)
    gtc.alignment = center
    gtc.border = thin
    ws.row_dimensions[total_row].height = 22

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"orders_{date_filter}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ── Order History ─────────────────────────────────────────────────────────────
@app.get("/orders/history")
def get_order_history(order_date: Optional[str] = None, db=Depends(get_db), token=Depends(verify_token)):
    cur = db.cursor(dictionary=True)
    user_id = int(token["sub"])
    role = token.get("role")
    date_filter = order_date or datetime.now().strftime("%Y-%m-%d")
    base_query = """
        SELECT h.id, h.dist_id, h.product_id,
               h.order_date, h.old_quantity, h.new_quantity, h.action, h.changed_at,
               d.name AS dist_name, d.code AS dist_code,
               p.name AS product_name, p.code AS product_code,
               u.name AS changed_by_name
        FROM order_history h
        JOIN distributors d ON h.dist_id = d.id
        JOIN products p ON h.product_id = p.id
        JOIN users u ON h.changed_by = u.id
        WHERE DATE(h.order_date) = %s
    """
    if role == "admin":
        cur.execute(base_query + " ORDER BY h.changed_at DESC", (date_filter,))
    else:
        cur.execute(base_query + " AND h.changed_by = %s ORDER BY h.changed_at DESC",
                    (date_filter, user_id))
    return cur.fetchall()

# ── Users / assignments ───────────────────────────────────────────────────────
@app.get("/users")
def list_users(db=Depends(get_db), _=Depends(require_admin)):
    cur = db.cursor(dictionary=True)
    cur.execute("SELECT id, name, email, role FROM users WHERE is_active=1 AND role='sales' ORDER BY name")
    return cur.fetchall()

@app.get("/users/{user_id}/distributors")
def get_user_distributors(user_id: int, db=Depends(get_db), _=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("SELECT dist_id FROM user_distributor_map WHERE user_id=%s", (user_id,))
    return [r[0] for r in cur.fetchall()]

@app.put("/users/{user_id}/distributors")
def set_user_distributors(user_id: int, data: UserDistributorAssignment,
                          db=Depends(get_db), _=Depends(require_admin)):
    cur = db.cursor()
    cur.execute("DELETE FROM user_distributor_map WHERE user_id=%s", (user_id,))
    for dist_id in data.dist_ids:
        cur.execute("INSERT IGNORE INTO user_distributor_map (user_id, dist_id) VALUES (%s, %s)",
                    (user_id, dist_id))
    db.commit()
    return {"message": f"{len(data.dist_ids)} distributors assigned"}
